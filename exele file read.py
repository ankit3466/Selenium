import openpyxl

path="C:/Users/user/Downloads/predicted_sheet1.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active   # if there are only one sheet in Exel file
#workbook.get_sheet_by_name("sheet_name")  # to choose a sheet from multiple sheet

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="    ")
    print()

